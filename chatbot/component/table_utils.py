class ExcelReadUtil(object):
    def __init__(self, excel_file, file_type):
        self.file_type = file_type
        self.header = []
        self.header_start_row = 1
        self.content_start_row = 2
        if self.file_type == 'xls':
            from xlrd import open_workbook
            self.work_book = open_workbook(file_contents=excel_file.read())
            self.sheet = self.work_book.worksheets()[0]
            self.row_num = self.sheet.nrows
            self.col_num = self.sheet.ncols

        elif self.file_type == 'xlsx':
            from openpyxl import load_workbook
            self.work_book = load_workbook(excel_file)
            self.sheet = self.work_book.worksheets[0]
            self.row_num = self.sheet.max_row
            self.col_num = self.sheet.max_column
            self.merged_cells = self.sheet.merged_cells
            self.header_end_col = self.col_num
            for index, rg in enumerate(self.merged_cells.ranges):
                start_col, start_row, end_col, end_row = rg.bounds
                # 如果第一行做了合并单元格，则从未合并的行开始提取表头
                if index == 0 and start_row == 1:
                    self.header_start_row = end_row + 1
                    self.header_end_col = end_col
                    self.content_start_row = self.header_start_row + 1
                    break
        else:
            pass

    def get_table_header(self):
        if self.header:
            return self.header.copy()
        if self.row_num > 1 and self.col_num > 1:
            if self.file_type == 'xls':
                self.header = self.sheet.row_values(0)
                return self.header.copy()
            else:

                for row in self.sheet.iter_rows(min_row=self.header_start_row, max_col=self.header_end_col,
                                                max_row=self.header_start_row + 1):
                    for cell in row:
                        self.header.append(str(cell.value))
                    return self.header.copy()
        else:
            return Exception("Excel格式不正确")

    def get_cell_value(self, row, col):
        if row > self.row_num:
            return Exception("超过最大行数")
        if col > self.col_num:
            return Exception("超过最大列数")
        if self.file_type == 'xls':
            return self.sheet.cell_value(row, col)

        if self.file_type == 'xlsx':
            return self.sheet.cell(row, col).value

    def get_cell_type(self, row, col):
        if row > self.row_num:
            return Exception("超过最大行数")
        if col > self.col_num:
            return Exception("超过最大列数")

        if self.file_type == 'xls':
            return self.sheet.cell_type(row, col)

        if self.file_type == 'xlsx':
            return self.sheet.cell(row, col).data_type


class ExcelWriteUtil(object):
    def __init__(self, filename, file_type='xlsx'):
        self.file_type = file_type
        self.filename = filename

        if file_type == 'xlsx':
            from openpyxl import Workbook
            self.excel_file = Workbook()
            self.work_sheet = self.excel_file.active
            self.work_sheet.title = 'sheet1'
            self.header = []

    def get_work_sheet(self):
        return self.work_sheet

    def get_work_book(self):
        return self.excel_file

    def save(self):
        self.excel_file.save(filename=self.filename)

    def set_header(self, headers):
        for idx, h in enumerate(headers):
            self.work_sheet.cell(row=1, column=idx + 1, value=h)
            self.header.append(h)

    def merge_header(self, headers):
        header_len = len(self.header)
        for h in headers:
            if h in self.header:
                continue
            header_len += 1
            self.work_sheet.cell(row=1, column=header_len, value=h)
            self.header.append(h)

    def append_row(self, row: dict):
        self.merge_header(row.keys())
        append_row = self.work_sheet.max_row + 1
        for idx, h in enumerate(self.header):
            self.work_sheet.cell(row=append_row, column=idx + 1, value=row.get(h, ''))

    def bulk_append_by_list_dict(self, rows: [{}]):
        for row in rows:
            self.append_row(row)
